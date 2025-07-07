import os
import cv2
import numpy as np
from datetime import datetime
from PIL import Image, ImageEnhance, ImageFilter
import pandas as pd
import pytesseract
from pdf2image import convert_from_path
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import tempfile
import shutil

class StreamlitOCREngine:
    def __init__(self):
        self.setup_cloud_environment()
        self.confidence_threshold = 10  # Lower for better detection
        self.row_height_threshold = 25
        self.word_spacing_threshold = 30  # Pixels for word grouping
        self.column_break_threshold = 80  # Pixels for column breaks
    
    def setup_cloud_environment(self):
        """Setup OCR for cloud deployment"""
        tesseract_paths = [
            '/usr/bin/tesseract',
            '/usr/local/bin/tesseract',
            '/opt/homebrew/bin/tesseract',
            r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        ]
        
        for path in tesseract_paths:
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd = path
                break
        
        tessdata_paths = [
            '/usr/share/tesseract-ocr/4.00/tessdata',
            '/usr/share/tesseract-ocr/tessdata',
            '/usr/local/share/tessdata'
        ]
        
        for path in tessdata_paths:
            if os.path.exists(path):
                os.environ['TESSDATA_PREFIX'] = path
                break
    
    def preprocess_image_for_forms(self, image):
        """Optimized preprocessing for form documents"""
        try:
            if isinstance(image, Image.Image):
                opencv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
            else:
                opencv_image = image
            
            gray = cv2.cvtColor(opencv_image, cv2.COLOR_BGR2GRAY)
            
            # Minimal preprocessing for clean documents
            denoised = cv2.fastNlMeansDenoising(gray, h=3, templateWindowSize=7, searchWindowSize=21)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(denoised)
            
            return Image.fromarray(enhanced)
            
        except Exception as e:
            print(f"Error in preprocessing: {str(e)}")
            return image
    
    def extract_text_with_word_level_data(self, image):
        """Extract text with word-level positioning data for smart spacing"""
        try:
            processed_image = self.preprocess_image_for_forms(image)
            
            # Try multiple OCR configurations for best results
            configs = [
                '--psm 6 -c preserve_interword_spaces=1',
                '--psm 4 -c preserve_interword_spaces=1',
                '--psm 3 -c preserve_interword_spaces=1',
                '--psm 11 -c preserve_interword_spaces=1'
            ]
            
            best_words = []
            best_score = 0
            
            for config in configs:
                try:
                    # Get word-level OCR data
                    ocr_data = pytesseract.image_to_data(
                        processed_image, 
                        output_type=pytesseract.Output.DICT,
                        config=config
                    )
                    
                    # Extract word-level elements
                    words = []
                    for i in range(len(ocr_data['text'])):
                        text = ocr_data['text'][i].strip()
                        confidence = int(ocr_data['conf'][i]) if ocr_data['conf'][i] != '-1' else 0
                        
                        if text and len(text.strip()) >= 1 and confidence >= self.confidence_threshold:
                            words.append({
                                'text': self.gentle_text_cleaning(text),
                                'left': ocr_data['left'][i],
                                'top': ocr_data['top'][i],
                                'width': ocr_data['width'][i],
                                'height': ocr_data['height'][i],
                                'confidence': confidence,
                                'right': ocr_data['left'][i] + ocr_data['width'][i]
                            })
                    
                    # Score this configuration
                    if words:
                        score = self.calculate_extraction_score(words)
                        if score > best_score:
                            best_score = score
                            best_words = words
                            
                except Exception as e:
                    print(f"OCR config {config} failed: {e}")
                    continue
            
            return best_words
            
        except Exception as e:
            print(f"Error extracting word-level data: {str(e)}")
            return []
    
    def calculate_extraction_score(self, words):
        """Calculate quality score for OCR results"""
        if not words:
            return 0
        
        total_confidence = sum(word['confidence'] for word in words)
        avg_confidence = total_confidence / len(words)
        total_text_length = sum(len(word['text']) for word in words)
        
        # Score based on quantity, confidence, and text length
        score = (len(words) * 0.4) + (avg_confidence * 0.3) + (total_text_length * 0.3)
        return score
    
    def group_words_by_spacing(self, words):
        """Group words based on horizontal spacing to preserve natural text flow"""
        if not words:
            return []
        
        # Sort words by vertical position first, then horizontal
        words.sort(key=lambda x: (x['top'], x['left']))
        
        # Group into rows
        rows = []
        current_row = []
        current_top = None
        
        for word in words:
            if current_top is None or abs(word['top'] - current_top) <= self.row_height_threshold:
                current_row.append(word)
                current_top = word['top'] if current_top is None else current_top
            else:
                if current_row:
                    # Process the completed row
                    processed_row = self.process_row_spacing(current_row)
                    if processed_row:
                        rows.append(processed_row)
                current_row = [word]
                current_top = word['top']
        
        # Process the last row
        if current_row:
            processed_row = self.process_row_spacing(current_row)
            if processed_row:
                rows.append(processed_row)
        
        return rows
    
    def process_row_spacing(self, row_words):
        """Process a single row to group words by spacing"""
        if not row_words:
            return []
        
        # Sort words in the row by horizontal position
        row_words.sort(key=lambda x: x['left'])
        
        # Group words based on spacing
        text_groups = []
        current_group = [row_words[0]]
        
        for i in range(1, len(row_words)):
            current_word = row_words[i]
            previous_word = row_words[i-1]
            
            # Calculate gap between words
            gap = current_word['left'] - previous_word['right']
            
            # Determine if words should be grouped together
            if gap <= self.word_spacing_threshold:
                # Small gap - keep words together
                current_group.append(current_word)
            else:
                # Large gap - start new group
                if current_group:
                    merged_group = self.merge_word_group(current_group)
                    if merged_group:
                        text_groups.append(merged_group)
                current_group = [current_word]
        
        # Add the last group
        if current_group:
            merged_group = self.merge_word_group(current_group)
            if merged_group:
                text_groups.append(merged_group)
        
        return text_groups
    
    def merge_word_group(self, word_group):
        """Merge a group of words into a single text element"""
        if not word_group:
            return None
        
        # Combine text with single spaces
        combined_text = ' '.join(word['text'] for word in word_group if word['text'].strip())
        
        if not combined_text.strip():
            return None
        
        # Calculate bounding box for the group
        left = min(word['left'] for word in word_group)
        top = min(word['top'] for word in word_group)
        right = max(word['right'] for word in word_group)
        bottom = max(word['top'] + word['height'] for word in word_group)
        
        # Calculate average confidence
        avg_confidence = sum(word['confidence'] for word in word_group) / len(word_group)
        
        return {
            'text': combined_text,
            'left': left,
            'top': top,
            'width': right - left,
            'height': bottom - top,
            'confidence': avg_confidence
        }
    
    def detect_column_breaks(self, grouped_rows):
        """Detect major column breaks in grouped text"""
        if not grouped_rows:
            return []
        
        enhanced_rows = []
        
        for row_groups in grouped_rows:
            if len(row_groups) <= 1:
                enhanced_rows.append(row_groups)
                continue
            
            # Analyze gaps between groups in this row
            final_groups = []
            
            for i, group in enumerate(row_groups):
                if i == 0:
                    final_groups.append(group)
                else:
                    # Check gap from previous group
                    prev_group = row_groups[i-1]
                    gap = group['left'] - (prev_group['left'] + prev_group['width'])
                    
                    # Add spacing indicator for large gaps
                    if gap >= self.column_break_threshold:
                        # Large gap detected - this helps with column placement
                        group['column_break'] = True
                    
                    final_groups.append(group)
            
            enhanced_rows.append(final_groups)
        
        return enhanced_rows
    
    def create_smart_excel_layout(self, grouped_rows, filename):
        """Create Excel with intelligent column placement based on spacing"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Smart_Layout"
            
            if not grouped_rows:
                return None
            
            excel_row = 1
            
            for row_groups in grouped_rows:
                if not row_groups:
                    excel_row += 1
                    continue
                
                # Determine column positions based on horizontal positions and breaks
                column_positions = self.calculate_smart_columns(row_groups)
                
                for group, col_pos in zip(row_groups, column_positions):
                    if group and group['text'].strip():
                        cell = ws.cell(row=excel_row, column=col_pos)
                        cell.value = group['text']
                        self.format_cell_smart(cell, group)
                
                excel_row += 1
            
            # Auto-adjust column widths
            self.adjust_column_widths_smart(ws)
            
            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            return temp_file.name
            
        except Exception as e:
            print(f"Error creating smart Excel layout: {str(e)}")
            return None
    
    def calculate_smart_columns(self, row_groups):
        """Calculate column positions based on horizontal positioning and breaks"""
        if not row_groups:
            return []
        
        column_positions = []
        current_column = 1
        
        for i, group in enumerate(row_groups):
            if i == 0:
                column_positions.append(current_column)
            else:
                # Check if this group has a column break marker
                if group.get('column_break', False):
                    current_column += 1
                
                column_positions.append(current_column)
        
        return column_positions
    
    def format_cell_smart(self, cell, element):
        """Apply smart formatting based on text characteristics"""
        text = element['text']
        
        # Set base alignment
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Format based on content patterns
        if text.isupper() and len(text) > 3:
            # Headers
            cell.font = Font(bold=True, size=12, color="1F4E79")
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        elif text.startswith(('•', '-', '*', '○', '▪')):
            # Bullet points
            cell.font = Font(size=10)
            cell.alignment = Alignment(indent=1, wrap_text=True, vertical='top')
        elif re.match(r'^\d+\.', text):
            # Numbered items
            cell.font = Font(bold=True, size=11, color="2E75B6")
        elif len(text) > 50:
            # Long text blocks
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='justify')
        else:
            # Regular text
            cell.font = Font(size=10)
    
    def adjust_column_widths_smart(self, worksheet):
        """Intelligently adjust column widths based on content"""
        column_widths = {}
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    column = cell.column_letter
                    content_length = len(str(cell.value))
                    
                    if column not in column_widths:
                        column_widths[column] = content_length
                    else:
                        column_widths[column] = max(column_widths[column], content_length)
        
        for column, width in column_widths.items():
            # Set reasonable width limits
            adjusted_width = min(max(width + 3, 15), 80)
            worksheet.column_dimensions[column].width = adjusted_width
    
    def gentle_text_cleaning(self, text):
        """Gentle text cleaning that preserves structure"""
        if not text:
            return ""
        
        # Only clean excessive whitespace
        text = re.sub(r'\s+', ' ', text.strip())
        
        # Fix only obvious OCR errors
        if '|' in text:
            text = text.replace('|', 'I')
        
        return text
    
    def process_image(self, image_file, filename):
        """Process image with smart spacing-based layout"""
        try:
            # Open image
            image = Image.open(image_file)
            
            # Extract words with positioning
            words = self.extract_text_with_word_level_data(image)
            
            if not words:
                return None, "No text detected. Try adjusting confidence threshold or spacing controls."
            
            # Group words by spacing
            grouped_rows = self.group_words_by_spacing(words)
            
            # Detect and handle column breaks
            enhanced_rows = self.detect_column_breaks(grouped_rows)
            
            # Create Excel with smart layout
            excel_path = self.create_smart_excel_layout(enhanced_rows, filename)
            
            if excel_path:
                total_groups = sum(len(row) for row in enhanced_rows)
                return excel_path, f"Extracted {len(words)} words grouped into {total_groups} text elements across {len(enhanced_rows)} rows"
            else:
                return None, "Failed to create Excel output"
                
        except Exception as e:
            return None, f"Error processing image: {str(e)}"
    
    def process_pdf(self, pdf_file, filename):
        """Process PDF with smart spacing-based layout"""
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_pdf_path = os.path.join(temp_dir, "temp.pdf")
                with open(temp_pdf_path, "wb") as f:
                    f.write(pdf_file.read())
                
                pages = convert_from_path(temp_pdf_path, dpi=300, fmt='png')
                
                excel_files = []
                total_words = 0
                
                for page_num, page in enumerate(pages, 1):
                    words = self.extract_text_with_word_level_data(page)
                    
                    if words:
                        grouped_rows = self.group_words_by_spacing(words)
                        enhanced_rows = self.detect_column_breaks(grouped_rows)
                        excel_path = self.create_smart_excel_layout(enhanced_rows, f"{filename}_page_{page_num}")
                        
                        if excel_path:
                            excel_files.append((excel_path, f"Page {page_num}"))
                            total_words += len(words)
                
                if excel_files:
                    return excel_files, f"Processed {len(pages)} pages with smart spacing, extracted {total_words} words"
                else:
                    return [], f"No text found in {len(pages)} pages"
                
        except Exception as e:
            return [], f"Error processing PDF: {str(e)}"
